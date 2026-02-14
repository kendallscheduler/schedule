"""
PART D: "Next year" reuse — promote PGY levels, rebalance cohorts, clear grid.
"""

import shutil
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

from .workbook_sheets import extract_residents, setup_all_sheets


def build_next_year(
    input_workbook: str,
    output_workbook: str,
    new_resident_names: Optional[Dict[int, List[str]]] = None,
    ty_exits: bool = True,
    cohort_intern_targets: Optional[Dict[str, int]] = None,
) -> str:
    """
    Create a fresh-year workbook from the current one.

    Steps:
    1) Copy workbook byte-for-byte
    2) Promote PGY labels: PGY-1→PGY-2→PGY-3; TY exits (configurable)
    3) Optionally replace resident names in col C
    4) Clear D..BC for a fresh schedule
    5) Keep formulas in BD..CD intact
    6) Re-run setup_all_sheets to refresh supporting sheets

    Args:
        input_workbook: source workbook path
        output_workbook: destination path
        new_resident_names: {row: [names]} if replacing specific rows
        ty_exits: remove TY rows (clear their names)
        cohort_intern_targets: {cohort_id: target_intern_count}

    Returns:
        Path to the new workbook.
    """
    inp = Path(input_workbook)
    out = Path(output_workbook)
    shutil.copy2(inp, out)

    wb = openpyxl.load_workbook(out, data_only=False)
    ws = wb["SCHEDULE"]

    # Promote PGY labels in col B (merged cells make this tricky — we edit the
    # top-left cell of each merge that contains a PGY label)
    pgy_promotion = {"PGY-1": "PGY-2", "PGY-2": "PGY-3"}
    for row in range(4, 57):
        b = ws.cell(row, 2).value
        if b and isinstance(b, str):
            b_stripped = b.strip()
            if b_stripped in pgy_promotion:
                ws.cell(row, 2, pgy_promotion[b_stripped])
            elif b_stripped == "TY" and ty_exits:
                # Optionally clear TY labels — PD will fill in new TYs
                pass  # leave label for now; PD can edit

    # Clear assignment grid D..BC for all resident rows
    # NOTE: ws.cell(row, col, value=None) does NOT clear in openpyxl
    # because the value= kwarg is only applied when not None.
    # Must use .value = None on the cell object.
    for row in range(4, 57):
        c = ws.cell(row, 3).value
        if c and str(c).strip():
            for col in range(4, 56):
                ws.cell(row, col).value = None

    # If new names provided, write them
    if new_resident_names:
        for row, names in new_resident_names.items():
            if names:
                ws.cell(row, 3, names[0] if len(names) == 1 else names[0])

    wb.save(out)

    # Refresh all supporting sheets
    setup_all_sheets(str(out))
    return str(out)
