"""
Write scheduled assignments to the Excel workbook.
Preserves template format, formulas, conditional formatting.
Adds CONFLICTS sheet when solver reports issues.
"""

import shutil
from pathlib import Path
from typing import Dict, List

import openpyxl


def write_schedule(
    template_path: str,
    output_path: str,
    assignments: Dict[str, Dict[int, str]],
    resident_row_map: Dict[str, int],
    week_start_col: int = 4,
    week_count: int = 52,
) -> str:
    """
    Copy template byte-for-byte, then write only D..BC values.
    Never touches styles, CF, merges, or formulas in BD..CD.
    """
    template = Path(template_path)
    output = Path(output_path)
    if not template.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    shutil.copy2(template, output)
    wb = openpyxl.load_workbook(output, data_only=False)
    ws = wb["SCHEDULE"]

    for name, week_data in assignments.items():
        row = resident_row_map.get(name)
        if row is None:
            continue
        for week, code in week_data.items():
            if 1 <= week <= week_count and code:
                col = week_start_col + (week - 1)
                ws.cell(row=row, column=col, value=code)

    wb.save(output)
    return str(output)


def add_conflicts_sheet(
    wb_path: str,
    conflicts: List[str],
) -> None:
    """Add a CONFLICTS sheet listing infeasibility messages."""
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    if "CONFLICTS" in wb.sheetnames:
        del wb["CONFLICTS"]
    ws = wb.create_sheet("CONFLICTS")
    ws.cell(1, 1, "Conflict / Issue")
    ws.cell(1, 2, "Suggestion")
    for i, msg in enumerate(conflicts, 2):
        ws.cell(i, 1, msg)
    wb.save(wb_path)
