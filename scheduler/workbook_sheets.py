"""
PART A + B: Add data-entry sheets and PROGRESS sheet to the workbook.
All sheets are added to the SAME workbook — SCHEDULE is never modified.
"""

import re
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Tuple

from .models import (
    Resident, Requirement, VacationRequest, CoverageRule, CohortDef,
    COUNTER_COLUMNS,
)


# ---------------------------------------------------------------------------
# Parse residents from the SCHEDULE sheet (source of truth)
# ---------------------------------------------------------------------------
def _parse_pgy(s):
    if not s or not isinstance(s, str):
        return None
    s = s.strip().upper()
    if s == "TY":
        return (1, True)
    m = re.match(r"PGY-?(\d)", s)
    if m:
        return (int(m.group(1)), False)
    return None


def extract_residents(ws) -> Tuple[List[Resident], Dict[str, int]]:
    """Parse SCHEDULE rows 4..56 → list of Resident + name→row map."""
    residents = []
    row_map = {}
    current_cohort = None
    current_pgy = None
    current_is_ty = False

    for row in range(4, 57):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        c = ws.cell(row, 3).value

        if a and str(a).strip().startswith("Cohort"):
            current_cohort = str(a).strip()
        if b:
            parsed = _parse_pgy(str(b))
            if parsed:
                current_pgy, current_is_ty = parsed

        if not c or not str(c).strip():
            continue
        if current_pgy is None:
            continue

        name = str(c).strip()
        r = Resident(
            name=name,
            pgy=current_pgy,
            is_ty=current_is_ty,
            cohort_id=current_cohort if not current_is_ty else None,
            row_index=row,
        )
        residents.append(r)
        row_map[name] = row

    return residents, row_map


# ---------------------------------------------------------------------------
# Derive clinic week patterns from current assignments
# ---------------------------------------------------------------------------
def derive_clinic_weeks(ws, residents: List[Resident]) -> Dict[str, List[int]]:
    """
    Return {cohort_id: sorted clinic week numbers} from existing data.
    Uses majority voting: a week counts as "clinic week" for the cohort
    if at least half the cohort's non-TY members are on clinic that week.
    """
    cohort_members: Dict[str, List[Resident]] = {}
    for res in residents:
        if res.cohort_id and not res.is_ty and res.row_index:
            cohort_members.setdefault(res.cohort_id, []).append(res)

    cohort_weeks: Dict[str, List[int]] = {}
    for coh_id, members in cohort_members.items():
        threshold = max(1, len(members) // 2)
        weeks = []
        for col in range(4, 56):
            week_num = col - 3
            count = 0
            for res in members:
                v = ws.cell(res.row_index, col).value
                if v and isinstance(v, str) and ("CLINIC" in v.upper() or "URGENT" in v.upper()):
                    count += 1
            if count >= threshold:
                weeks.append(week_num)
        cohort_weeks[coh_id] = weeks
    return cohort_weeks


# ---------------------------------------------------------------------------
# Default requirement targets (from PD program rules)
# ---------------------------------------------------------------------------
def default_requirements() -> List[Requirement]:
    return [
        # PGY1
        Requirement("PGY1", "ICU",      8, notes="Day + Night combined"),
        Requirement("PGY1", "FLOORS",   20, notes="A/B/C/D/G + NF +/- Swing"),
        Requirement("PGY1", "CLINIC",   14),
        Requirement("PGY1", "CARDIO",   2),
        Requirement("PGY1", "ID",       2),
        Requirement("PGY1", "ED",       2),
        Requirement("PGY1", "VACATION", 4),
        # PGY2
        Requirement("PGY2", "ICU",      8),
        Requirement("PGY2", "FLOORS",   16),
        Requirement("PGY2", "CLINIC",   14),
        Requirement("PGY2", "CARDIO",   4),
        Requirement("PGY2", "ID",       2),
        Requirement("PGY2", "ED",       2),
        Requirement("PGY2", "NEURO",    2),
        Requirement("PGY2", "VACATION", 4),
        # PGY3 (lighter)
        Requirement("PGY3", "FLOORS",   16),
        Requirement("PGY3", "CLINIC",   14),
        Requirement("PGY3", "VACATION", 4),
        # TY (same as PGY1 for core)
        Requirement("TY",   "ICU",      8),
        Requirement("TY",   "FLOORS",   20),
        Requirement("TY",   "CLINIC",   4, notes="TY clinic lighter"),
        Requirement("TY",   "ID",       2),
        Requirement("TY",   "ED",       2),
        Requirement("TY",   "VACATION", 4),
    ]


# ---------------------------------------------------------------------------
# Default coverage rules
# ---------------------------------------------------------------------------
def default_coverage_rules() -> List[CoverageRule]:
    return [
        CoverageRule("FLOOR_A", 1, 1, 2, "1 senior + 2 interns per week"),
        CoverageRule("FLOOR_B", 1, 1, 2, "1 senior + 2 interns per week"),
        CoverageRule("FLOOR_C", 1, 1, 2, "1 senior + 2 interns per week"),
        CoverageRule("FLOOR_D", 1, 1, 2, "1 senior + 2 interns per week"),
        CoverageRule("ICU_DAY", 1, 2, 2, "2 seniors + 2 interns per week"),
        CoverageRule("NF",      1, 1, 1, "1 senior + 1 intern per week"),
        CoverageRule("ICU_NIGHT", 1, 1, 1, "1 senior + 1 intern per week"),
        CoverageRule("SWING",   1, 1, 0, "1 senior, no intern"),
        CoverageRule("TEAM_G",  0, 1, 0, "Optional; seniors only; 0 required"),
    ]


# ===================================================================
# PART A: Add data-entry sheets
# ===================================================================

def ensure_residents_sheet(wb, residents: List[Resident]):
    """Create/replace RESIDENTS sheet extracted from SCHEDULE."""
    if "RESIDENTS" in wb.sheetnames:
        del wb["RESIDENTS"]
    ws = wb.create_sheet("RESIDENTS")
    headers = ["ResidentName", "PGY", "IsSenior", "IsIntern", "IsTY",
               "CohortId", "Track", "Notes", "ScheduleRow"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for i, r in enumerate(residents, 2):
        pgy_label = "TY" if r.is_ty else f"PGY{r.pgy}"
        ws.cell(i, 1, r.name)
        ws.cell(i, 2, pgy_label)
        ws.cell(i, 3, "Y" if r.is_senior else "N")
        ws.cell(i, 4, "Y" if r.is_intern else "N")
        ws.cell(i, 5, "Y" if r.is_ty else "N")
        ws.cell(i, 6, r.cohort_id or "")
        ws.cell(i, 7, r.track or "")
        ws.cell(i, 8, r.notes or "")
        ws.cell(i, 9, r.row_index)


def ensure_requirements_sheet(wb, requirements: Optional[List[Requirement]] = None):
    """Create/replace REQUIREMENTS_TARGETS sheet."""
    if "REQUIREMENTS_TARGETS" not in wb.sheetnames:
        del wb["REQUIREMENTS_TARGETS"]
    else:
        # If it exists, let's keep it but check headers
        pass
    
    # Actually, for simplicity in this version, let's always recreate it if empty 
    # but the tool usually deletes and recreated. 
    # Re-using the same logic to ensure clean headers:
    if "REQUIREMENTS_TARGETS" in wb.sheetnames:
        del wb["REQUIREMENTS_TARGETS"]
    ws = wb.create_sheet("REQUIREMENTS_TARGETS")

    headers = ["PGY", "Category", "RequiredWeeks", "MinWeeks", "MaxWeeks", "IsMandatory", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    reqs = requirements or default_requirements()
    for i, req in enumerate(reqs, 2):
        ws.cell(i, 1, req.pgy)
        ws.cell(i, 2, req.category)
        ws.cell(i, 3, req.required_weeks)
        ws.cell(i, 4, req.min_weeks)
        ws.cell(i, 5, req.max_weeks)
        ws.cell(i, 6, "Y" if req.is_mandatory else "N")
        ws.cell(i, 7, req.notes or "")


def ensure_vacation_sheet(wb, vacations: Optional[List[VacationRequest]] = None):
    """Create/replace VACATION_REQUESTS sheet with headers (and any existing data)."""
    if "VACATION_REQUESTS" in wb.sheetnames:
        del wb["VACATION_REQUESTS"]
    ws = wb.create_sheet("VACATION_REQUESTS")
    headers = ["ResidentName", "PGY", "RequestType", "StartWeekNumber",
               "LengthWeeks", "Priority", "HardLock", "Comments"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    if vacations:
        for i, v in enumerate(vacations, 2):
            ws.cell(i, 1, v.resident_name)
            ws.cell(i, 2, v.pgy)
            ws.cell(i, 3, v.request_type)
            ws.cell(i, 4, v.start_week)
            ws.cell(i, 5, v.length_weeks)
            ws.cell(i, 6, v.priority)
            ws.cell(i, 7, "Y" if v.hard_lock else "N")
            ws.cell(i, 8, v.comments or "")


def ensure_coverage_sheet(wb, rules: Optional[List[CoverageRule]] = None):
    """Create/replace COVERAGE_RULES sheet."""
    if "COVERAGE_RULES" in wb.sheetnames:
        del wb["COVERAGE_RULES"]
    ws = wb.create_sheet("COVERAGE_RULES")
    headers = ["RotationPool", "RequiredPerWeek", "SeniorNeededPerUnit",
               "InternNeededPerUnit", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    cov = rules or default_coverage_rules()
    for i, rule in enumerate(cov, 2):
        ws.cell(i, 1, rule.rotation_pool)
        ws.cell(i, 2, rule.required_per_week)
        ws.cell(i, 3, rule.senior_per_unit)
        ws.cell(i, 4, rule.intern_per_unit)
        ws.cell(i, 5, rule.notes or "")


def ensure_cohorts_sheet(wb, cohort_defs: Optional[List[CohortDef]] = None,
                         clinic_weeks_map: Optional[Dict[str, List[int]]] = None):
    """Create/replace COHORTS sheet."""
    if "COHORTS" in wb.sheetnames:
        del wb["COHORTS"]
    ws = wb.create_sheet("COHORTS")
    headers = ["CohortId", "ClinicWeekPattern", "TargetInternCount", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    if cohort_defs:
        for i, cd in enumerate(cohort_defs, 2):
            ws.cell(i, 1, cd.cohort_id)
            ws.cell(i, 2, ",".join(str(w) for w in cd.clinic_weeks))
            ws.cell(i, 3, cd.target_intern_count)
            ws.cell(i, 4, cd.notes or "")
    elif clinic_weeks_map:
        i = 2
        for coh_id in sorted(clinic_weeks_map.keys()):
            ws.cell(i, 1, coh_id)
            ws.cell(i, 2, ",".join(str(w) for w in clinic_weeks_map[coh_id]))
            ws.cell(i, 3, "")
            ws.cell(i, 4, "Derived from current schedule")
            i += 1


def ensure_config_sheet(wb):
    """Create/replace SOLVER_CONFIG sheet with global params."""
    if "SOLVER_CONFIG" in wb.sheetnames:
        # Don't overwrite if it exists, to preserve user edits
        return 
    ws = wb.create_sheet("SOLVER_CONFIG")
    headers = ["Parameter", "Value", "Description"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    
    defaults = [
        ("max_nights_per_year", 8, "Max night shifts (ICU N, NF) per resident per year"),
        ("max_consecutive_nights", 4, "Max consecutive weeks on night shifts"),
        ("ed_cap_per_week", 3, "Max residents in ED per week"),
        ("ramirez_forbidden_until_week", 7, "No PGY1 on CARDIO-RAM before this week"),
        ("july_weeks", "1,2,3,4", "Weeks considered July (comma separated)"),
        ("vacation_weeks_per_resident", 4, "Total vacation weeks required per resident"),
        ("min_working_residents", 21, "Min residents needed for full coverage (warning only)"),
    ]
    for i, (p, v, d) in enumerate(defaults, 2):
        ws.cell(i, 1, p)
        ws.cell(i, 2, v)
        ws.cell(i, 3, d)


# ===================================================================
# PART B: PROGRESS sheet — links to BD..CD formulas
# ===================================================================

def ensure_progress_sheet(wb, residents: List[Resident],
                          requirements: Optional[List[Requirement]] = None):
    """
    Create/replace PROGRESS sheet.
    For each resident × requirement category, adds a cell-reference formula
    that reads the already-computed counter from SCHEDULE BD..CD.
    """
    if "PROGRESS" in wb.sheetnames:
        del wb["PROGRESS"]
    ws = wb.create_sheet("PROGRESS")

    headers = ["Resident", "PGY", "ScheduleRow",
               "Category", "Required", "Completed", "Remaining", "Flags"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)

    # Map category → counter column number
    cat_to_col: Dict[str, int] = {}
    for col_num, hdr in COUNTER_COLUMNS.items():
        if hdr not in ("SUB-TOTAL", "SUB-TOTAL2", "TOTAL"):
            cat_to_col[hdr] = col_num

    reqs = requirements or default_requirements()
    row = 2
    for res in residents:
        pgy_label = "TY" if res.is_ty else f"PGY{res.pgy}"
        matching_reqs = [r for r in reqs if r.pgy == pgy_label]
        if not matching_reqs:
            continue
        sched_row = res.row_index
        for req in matching_reqs:
            cat = req.category
            counter_col = cat_to_col.get(cat)

            ws.cell(row, 1, res.name)
            ws.cell(row, 2, pgy_label)
            ws.cell(row, 3, sched_row)
            ws.cell(row, 4, cat)
            ws.cell(row, 5, req.required_weeks)

            # "Completed" = link to the SCHEDULE counter cell
            if counter_col and sched_row:
                col_letter = get_column_letter(counter_col)
                ws.cell(row, 6).value = f"=SCHEDULE!{col_letter}{sched_row}"
            else:
                ws.cell(row, 6, 0)

            # "Remaining" = Required - Completed (min 0)
            ws.cell(row, 7).value = f"=MAX(0, E{row}-F{row})"

            # "Flags" = DEFICIT if Remaining > 0
            ws.cell(row, 8).value = f'=IF(G{row}>0,"DEFICIT","")'

            row += 1


# ===================================================================
# Master function: add ALL sheets to a workbook
# ===================================================================

def setup_all_sheets(wb_path: str, save: bool = True) -> str:
    """
    Open the workbook, extract residents, add/refresh all supporting sheets.
    Returns the path to the saved workbook.
    """
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    ws_sched = wb["SCHEDULE"]

    residents, row_map = extract_residents(ws_sched)
    clinic_weeks = derive_clinic_weeks(ws_sched, residents)

    ensure_residents_sheet(wb, residents)
    ensure_requirements_sheet(wb)
    ensure_vacation_sheet(wb)
    ensure_coverage_sheet(wb)
    ensure_cohorts_sheet(wb, clinic_weeks_map=clinic_weeks)
    ensure_config_sheet(wb)
    ensure_progress_sheet(wb, residents)

    if save:
        wb.save(wb_path)
    return wb_path
