"""
Parse inputs for the AutoScheduler — reads from the workbook's own sheets.
The workbook IS the source of truth: SCHEDULE + RESIDENTS + REQUIREMENTS_TARGETS +
VACATION_REQUESTS + COVERAGE_RULES + COHORTS are all in the same .xlsx.
"""

from typing import List, Optional

import openpyxl

from .models import (
    ScheduleContext, Resident, Requirement, VacationRequest,
    CoverageRule, CohortDef,
)
from .workbook_sheets import (
    extract_residents, default_requirements, default_coverage_rules,
)


def _read_requirements(wb) -> List[Requirement]:
    """Read REQUIREMENTS_TARGETS sheet; fall back to defaults."""
    if "REQUIREMENTS_TARGETS" not in wb.sheetnames:
        return default_requirements()
    ws = wb["REQUIREMENTS_TARGETS"]
    reqs = []
    for row in range(2, ws.max_row + 1):
        pgy = ws.cell(row, 1).value
        cat = ws.cell(row, 2).value
        req_w = ws.cell(row, 3).value
        if not pgy or not cat:
            continue
        
        mandatory_val = ws.cell(row, 6).value
        is_mandatory = str(mandatory_val).strip().upper().startswith("Y") if mandatory_val else False

        reqs.append(Requirement(
            pgy=str(pgy).strip(),
            category=str(cat).strip(),
            required_weeks=int(req_w) if req_w else 0,
            min_weeks=ws.cell(row, 4).value,
            max_weeks=ws.cell(row, 5).value,
            is_mandatory=is_mandatory,
            notes=ws.cell(row, 7).value,
        ))
    return reqs if reqs else default_requirements()


def _read_vacations(wb) -> List[VacationRequest]:
    """Read VACATION_REQUESTS sheet."""
    if "VACATION_REQUESTS" not in wb.sheetnames:
        return []
    ws = wb["VACATION_REQUESTS"]
    vacs = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name or not str(name).strip():
            continue
        vacs.append(VacationRequest(
            resident_name=str(name).strip(),
            pgy=str(ws.cell(row, 2).value or "").strip(),
            request_type=str(ws.cell(row, 3).value or "VAC_BLOCK_1").strip(),
            start_week=int(ws.cell(row, 4).value or 1),
            length_weeks=int(ws.cell(row, 5).value or 2),
            priority=int(ws.cell(row, 6).value or 3),
            hard_lock=str(ws.cell(row, 7).value or "N").strip().upper().startswith("Y"),
            comments=ws.cell(row, 8).value,
        ))
    return vacs


def _read_coverage(wb) -> List[CoverageRule]:
    """Read COVERAGE_RULES sheet; fall back to defaults."""
    if "COVERAGE_RULES" not in wb.sheetnames:
        return default_coverage_rules()
    ws = wb["COVERAGE_RULES"]
    rules = []
    for row in range(2, ws.max_row + 1):
        pool = ws.cell(row, 1).value
        if not pool:
            continue
        rules.append(CoverageRule(
            rotation_pool=str(pool).strip(),
            required_per_week=int(ws.cell(row, 2).value or 0),
            senior_per_unit=int(ws.cell(row, 3).value or 0),
            intern_per_unit=int(ws.cell(row, 4).value or 0),
            notes=ws.cell(row, 5).value,
        ))
    return rules if rules else default_coverage_rules()


def _read_cohorts(wb) -> List[CohortDef]:
    """Read COHORTS sheet."""
    if "COHORTS" not in wb.sheetnames:
        return []
    ws = wb["COHORTS"]
    defs = []
    for row in range(2, ws.max_row + 1):
        cid = ws.cell(row, 1).value
        if not cid:
            continue
        pattern_str = str(ws.cell(row, 2).value or "")
        weeks = []
        if pattern_str.strip():
            weeks = [int(x.strip()) for x in pattern_str.split(",") if x.strip().isdigit()]
        target = ws.cell(row, 3).value
        defs.append(CohortDef(
            cohort_id=str(cid).strip(),
            clinic_weeks=weeks,
            target_intern_count=int(target) if target else 2,
            notes=ws.cell(row, 4).value,
        ))
    return defs


from .models import SolverConfig

def _read_config(wb) -> SolverConfig:
    """Read SOLVER_CONFIG sheet."""
    config = SolverConfig()
    if "SOLVER_CONFIG" not in wb.sheetnames:
        return config
    
    ws = wb["SOLVER_CONFIG"]
    rows = {}
    for row in range(2, ws.max_row + 1):
        p = ws.cell(row, 1).value
        v = ws.cell(row, 2).value
        if p:
            rows[str(p).strip()] = v
            
    if "max_nights_per_year" in rows: config.max_nights_per_year = int(rows["max_nights_per_year"])
    if "max_consecutive_nights" in rows: config.max_consecutive_nights = int(rows["max_consecutive_nights"])
    if "ed_cap_per_week" in rows: config.ed_cap_per_week = int(rows["ed_cap_per_week"])
    if "ramirez_forbidden_until_week" in rows: config.ramirez_forbidden_until_week = int(rows["ramirez_forbidden_until_week"])
    if "july_weeks" in rows:
        j_str = str(rows["july_weeks"])
        config.july_weeks = [int(x.strip()) for x in j_str.split(",") if x.strip().isdigit()]
    if "vacation_weeks_per_resident" in rows: config.vacation_weeks_per_resident = int(rows["vacation_weeks_per_resident"])
    if "min_working_residents" in rows: config.min_working_residents = int(rows["min_working_residents"])
    
    return config


def parse_workbook(wb_path: str, random_seed: Optional[int] = None) -> ScheduleContext:
    """
    Parse the single workbook to build the full ScheduleContext.
    The workbook must already have had setup_all_sheets() called.
    """
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    ws_sched = wb["SCHEDULE"]

    residents, row_map = extract_residents(ws_sched)
    requirements = _read_requirements(wb)
    vacations = _read_vacations(wb)
    coverage = _read_coverage(wb)
    cohorts = _read_cohorts(wb)
    config = _read_config(wb)

    return ScheduleContext(
        residents=residents,
        requirements=requirements,
        vacation_requests=vacations,
        coverage_rules=coverage,
        cohort_defs=cohorts,
        config=config,
        resident_row_map=row_map,
        week_count=52,
        random_seed=random_seed,
    )
