
import openpyxl

def audit_requirements(path):
    print(f"Auditing Requirements in: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["SCHEDULE"]
    
    current_pgy = None
    for r in range(4, 57):
        pgy_val = ws.cell(r, 2).value
        if pgy_val:
            current_pgy = str(pgy_val).strip()
        
        name = ws.cell(r, 3).value
        if not name: continue
        
        counts = {}
        for c in range(4, 56): # D..BC
            val = ws.cell(r, c).value
            if val:
                v = str(val).strip().upper()
                counts[v] = counts.get(v, 0) + 1
        
        # Floor Category: A, B, C, D, G, NF, SWING
        floor_sum = sum(counts.get(k, 0) for k in ['A', 'B', 'C', 'D', 'G', 'NF', 'SWING'])
        clinic_sum = sum(counts.get(k, 0) for k in ['CLINIC', 'CLINIC *', 'CLINIC*', 'URGENT CARE', 'URGENT CARE *'])
        icu_sum = sum(counts.get(k, 0) for k in ['ICU', 'ICU E', 'ICU H', 'ICU H*', 'ICU N'])
        
        print(f"{name:20} [{current_pgy:6}] | Floors: {floor_sum:2} | Clinic: {clinic_sum:2} | ICU: {icu_sum:2}")

if __name__ == "__main__":
    audit_requirements("2025 Updated Schedule - Copy 2026.xlsx")
