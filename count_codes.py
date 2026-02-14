
import openpyxl

def count_resident_codes(path):
    print(f"Counting codes in: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["SCHEDULE"]
    
    for r in range(4, 15):
        name = ws.cell(r, 3).value
        pgy = ws.cell(r, 2).value
        if not name: continue
        
        counts = {}
        for c in range(4, 56):
            val = ws.cell(r, c).value
            if val:
                v = str(val).strip()
                counts[v] = counts.get(v, 0) + 1
        
        # Calculate floor sum (A, B, C, D, G)
        floor_sum = counts.get('A',0) + counts.get('B',0) + counts.get('C',0) + counts.get('D',0) + counts.get('G',0)
        
        print(f"{name:20} ({pgy}): FLOOR={floor_sum}, {counts}")

if __name__ == "__main__":
    count_resident_codes("2025 Updated Schedule - Copy 2026.xlsx")
