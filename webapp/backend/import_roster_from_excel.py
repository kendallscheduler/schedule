#!/usr/bin/env python3
"""Import residents from existing master schedule Excel into the webapp DB."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from database import SessionLocal
from models import Resident, Year, Cohort

def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else "../../2025 Updated Schedule - Copy 2026.xlsx"
    excel_path = Path(__file__).resolve().parent / excel_path
    if not excel_path.exists():
        print(f"File not found: {excel_path}")
        sys.exit(1)

    db = SessionLocal()
    year = db.query(Year).filter(Year.name == "2025-2026").first()
    if not year:
        year = Year(name="2025-2026", start_date="2025-07-01")
        db.add(year)
        db.commit()
        db.refresh(year)

    cohorts = {c.name: c for c in db.query(Cohort).filter(Cohort.year_id == year.id).all()}
    if not cohorts:
        for i, name in enumerate(["Cohort 1", "Cohort 2", "Cohort 3", "Cohort 4", "Cohort 5"], 1):
            c = Cohort(year_id=year.id, name=name, clinic_weeks=[], target_intern_count=2)
            db.add(c)
        db.commit()
        cohorts = {c.name: c for c in db.query(Cohort).filter(Cohort.year_id == year.id).all()}

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["SCHEDULE"]
    current_cohort = None
    current_pgy = None
    created = 0

    for row in range(4, 57):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        c = ws.cell(row, 3).value
        if a and str(a).strip().startswith("Cohort"):
            current_cohort = str(a).strip()
        if b:
            pgy_str = str(b).strip().upper()
            if pgy_str in ("PGY-1", "PGY-2", "PGY-3"):
                current_pgy = pgy_str.replace("-", "")
            elif pgy_str == "TY":
                current_pgy = "TY"
        if not c or not str(c).strip():
            continue
        if not current_pgy:
            continue

        name = str(c).strip()
        cohort_id = cohorts.get(current_cohort).id if current_cohort and current_cohort in cohorts else None
        if current_pgy == "PGY1":
            constraints = {"no_cardio_before_week": 7}
        else:
            constraints = {}

        existing = db.query(Resident).filter(Resident.name == name, Resident.year_id == year.id).first()
        if not existing:
            r = Resident(name=name, pgy=current_pgy, cohort_id=cohort_id, year_id=year.id, constraints_json=constraints)
            db.add(r)
            created += 1

    db.commit()
    print(f"Imported {created} residents. Total: {db.query(Resident).filter(Resident.year_id == year.id).count()}")
    db.close()

if __name__ == "__main__":
    main()
