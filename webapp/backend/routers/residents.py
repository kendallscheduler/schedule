from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
import pandas as pd
import openpyxl
from io import BytesIO

from database import get_db
from models import Resident, Year, Cohort, ScheduleAssignment
from schemas import ResidentCreate, ResidentUpdate, ResidentOut, PasteScheduleRequest
from engine import MAX_COHORT_SIZE

router = APIRouter()


def _is_intern(pgy: str) -> bool:
    return pgy in ("PGY1", "TY")


def _check_cohort_size(db: Session, year_id: int, cohort_id: Optional[int], exclude_resident_id: Optional[int] = None) -> None:
    """Raise HTTPException if assigning to this cohort would exceed MAX_COHORT_SIZE."""
    if not cohort_id:
        return
    count = db.query(Resident).filter(Resident.year_id == year_id, Resident.cohort_id == cohort_id).count()
    if exclude_resident_id:
        r = db.query(Resident).filter(Resident.id == exclude_resident_id).first()
        if r and r.cohort_id == cohort_id:
            count -= 1
    if count >= MAX_COHORT_SIZE:
        c = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        cname = c.name if c else str(cohort_id)
        raise HTTPException(400, f"Cohort {cname} already has {MAX_COHORT_SIZE} residents. Max per cohort is {MAX_COHORT_SIZE}.")


def _check_interns_even(db: Session, year_id: int, cohort_id: Optional[int], pgy: str, exclude_resident_id: Optional[int] = None) -> None:
    """Interns per cohort must be in multiples of 2 (each needs a co-intern). Raise if odd."""
    if not cohort_id or not _is_intern(pgy):
        return
    q = db.query(Resident).filter(Resident.year_id == year_id, Resident.cohort_id == cohort_id)
    if exclude_resident_id:
        r = db.query(Resident).filter(Resident.id == exclude_resident_id).first()
        if r and r.cohort_id == cohort_id and _is_intern(r.pgy):
            q = q.filter(Resident.id != exclude_resident_id)
    intern_count = sum(1 for res in q.all() if _is_intern(res.pgy))
    after_count = intern_count + 1
    if after_count % 2 != 0:
        c = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        cname = c.name if c else str(cohort_id)
        raise HTTPException(400, f"Cohort {cname} would have {after_count} interns. Interns must be in multiples of 2 (each needs a co-intern). Add or remove 1 intern.")

# Map Excel rotation values to canonical codes (for schedule import)
ROT_NORMALIZE = {
    "a": "A", "b": "B", "c": "C", "d": "D", "g": "G",
    "icu": "ICU", "icu e": "ICU E", "icu-e": "ICU E", "icu n": "ICU N", "icun": "ICU N",
    "nf": "NF", "swing": "SWING",
    "clinic": "CLINIC", "clinic *": "CLINIC *", "clinic*": "CLINIC *",
    "ed": "ED", "cardio": "CARDIO", "cardio-ram": "CARDIO-RAM", "cardio-hca": "CARDIO-HCA",
    "id": "ID", "neuro": "NEURO", "vacation": "VACATION", "vac": "VACATION",
    "geriatrics": "GERIATRICS", "geri": "GERIATRICS",
    "pulmonology": "PULMONOLOGY", "pulm": "PULMONOLOGY",
    "nephrology": "NEPHROLOGY", "nephro": "NEPHROLOGY",
    "palliative": "PALLIATIVE", "pain": "PAIN",
    "rheumatology": "RHEUMATOLOGY", "rheum": "RHEUMATOLOGY",
    "endocrinology": "ENDOCRINOLOGY", "endo": "ENDOCRINOLOGY",
    "trauma": "TRAUMA", "sicu": "SICU", "plastic": "PLASTIC",
    "elective": "ELECTIVE",
    "icu h": "ICU H", "icu h*": "ICU H", "icu h *": "ICU H", "icuh": "ICU H", "icuh*": "ICU H",
    "cardio*": "CARDIO-RAM", "cardio *": "CARDIO-RAM",
    "id *": "ID", "id*": "ID",
}


@router.get("/", response_model=list[ResidentOut])
def list_residents(year_id: Optional[int] = None, cohort_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(Resident)
    if year_id:
        q = q.filter(Resident.year_id == year_id)
    if cohort_id is not None:
        q = q.filter(Resident.cohort_id == cohort_id)
    residents = q.all()
    return [ResidentOut.model_validate(r) for r in residents]


# NF counts as FLOORS. SWING counts as NF or ICU_NIGHT (balanced).
_FLOOR_ROTS = ["A", "B", "C", "D", "G"]
_ICU_ROTS = ["ICU", "ICU E", "ICU N"]
_ELECTIVE_ROTS = ["ELECTIVE", "CARDIO", "CARDIO-RAM", "CARDIO-HCA", "ID", "NEURO", "GERIATRICS",
                  "PULMONOLOGY", "NEPHROLOGY", "PALLIATIVE", "PAIN", "RHEUMATOLOGY", "ENDOCRINOLOGY", "ICU H"]


@router.get("/{resident_id}/rotation-history")
def get_rotation_history(resident_id: int, db: Session = Depends(get_db)):
    """Walk prior_resident_id chain and return past years' schedule assignments for reference."""
    r = db.query(Resident).filter(Resident.id == resident_id).first()
    if not r:
        raise HTTPException(404, "Resident not found")
    out = []
    cur = r
    seen = set()
    while cur and cur.prior_resident_id and cur.prior_resident_id not in seen:
        seen.add(cur.prior_resident_id)
        prior = db.query(Resident).filter(Resident.id == cur.prior_resident_id).first()
        if not prior:
            break
        year = db.query(Year).filter(Year.id == prior.year_id).first()
        year_name = year.name if year else str(prior.year_id)
        assigns = db.query(ScheduleAssignment).filter(
            ScheduleAssignment.resident_id == prior.id,
            ScheduleAssignment.year_id == prior.year_id,
        ).order_by(ScheduleAssignment.week_number).all()
        floor_cnt = sum(1 for a in assigns if a.rotation_code in _FLOOR_ROTS)
        nf_cnt = sum(1 for a in assigns if a.rotation_code == "NF")
        icun_cnt = sum(1 for a in assigns if a.rotation_code == "ICU N")
        swing_cnt = sum(1 for a in assigns if a.rotation_code == "SWING")
        x = (icun_cnt - nf_cnt + swing_cnt) // 2 if swing_cnt else 0
        swing_to_nf = max(0, min(swing_cnt, x))
        clinic_cnt = sum(1 for a in assigns if a.rotation_code in ("CLINIC", "CLINIC *"))
        ed_cnt = sum(1 for a in assigns if a.rotation_code == "ED")
        trauma_cnt = sum(1 for a in assigns if a.rotation_code == "TRAUMA")
        sicu_cnt = sum(1 for a in assigns if a.rotation_code == "SICU")
        plastic_cnt = sum(1 for a in assigns if a.rotation_code == "PLASTIC")
        elective_cnt = sum(1 for a in assigns if a.rotation_code in _ELECTIVE_ROTS)
        clinic_req = 14 if prior.pgy in ("PGY1", "PGY2", "PGY3") else 0
        elective_cnt += max(0, clinic_cnt - clinic_req)  # clinic overflow counts as elective
        by_cat = {
            "FLOORS": floor_cnt + nf_cnt + swing_to_nf,
            "ICU": sum(1 for a in assigns if a.rotation_code in _ICU_ROTS),
            "NF": nf_cnt + swing_to_nf,
            "ICU_NIGHT": icun_cnt + (swing_cnt - swing_to_nf),
            "SWING": swing_cnt,
            "CLINIC": clinic_cnt,
            "ED": ed_cnt,
            "TRAUMA": trauma_cnt,
            "SICU": sicu_cnt,
            "PLASTIC": plastic_cnt,
            "ELECTIVE": elective_cnt,
        }
        out.append({
            "year_name": year_name,
            "pgy": prior.pgy,
            "assignments": [{"week": a.week_number, "rotation": a.rotation_code} for a in assigns],
            "by_category": by_cat,
        })
        cur = prior
    return {"history": out}


@router.get("/{resident_id}", response_model=ResidentOut)
def get_resident(resident_id: int, db: Session = Depends(get_db)):
    r = db.query(Resident).filter(Resident.id == resident_id).first()
    if not r:
        raise HTTPException(404, "Resident not found")
    return ResidentOut.model_validate(r)


@router.post("/", response_model=ResidentOut)
def create_resident(data: ResidentCreate, db: Session = Depends(get_db)):
    _check_cohort_size(db, data.year_id, data.cohort_id)
    _check_interns_even(db, data.year_id, data.cohort_id, data.pgy)
    r = Resident(**data.model_dump())
    db.add(r)
    db.commit()
    db.refresh(r)
    return ResidentOut.model_validate(r)


@router.patch("/{resident_id}", response_model=ResidentOut)
def update_resident(resident_id: int, data: ResidentUpdate, db: Session = Depends(get_db)):
    r = db.query(Resident).filter(Resident.id == resident_id).first()
    if not r:
        raise HTTPException(404, "Resident not found")
    if "cohort_id" in data.model_dump(exclude_unset=True):
        new_cohort = data.cohort_id
        _check_cohort_size(db, r.year_id, new_cohort, exclude_resident_id=resident_id)
        new_pgy = data.pgy if data.pgy is not None else r.pgy
        _check_interns_even(db, r.year_id, new_cohort, new_pgy, exclude_resident_id=resident_id)
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    return ResidentOut.model_validate(r)


@router.delete("/{resident_id}")
def delete_resident(resident_id: int, db: Session = Depends(get_db)):
    r = db.query(Resident).filter(Resident.id == resident_id).first()
    if not r:
        raise HTTPException(404, "Resident not found")
    db.delete(r)
    db.commit()
    return {"ok": True}


@router.post("/{resident_id}/schedule-assignments")
def paste_schedule_assignments(resident_id: int, data: PasteScheduleRequest, db: Session = Depends(get_db)):
    """Paste schedule: whitespace-separated rotation codes. Skips leading Cohort/PGY/Name tokens."""
    r = db.query(Resident).filter(Resident.id == resident_id).first()
    if not r:
        raise HTTPException(404, "Resident not found")
    year_id = data.year_id
    tokens = data.assignments.split()
    # Find first token that normalizes to a valid rotation
    start = 0
    for i, t in enumerate(tokens):
        code = _norm_rot(t)
        if code:
            start = i
            break
    # Build assignments. Handle "CLINIC *" style: "*" as suffix to previous rotation.
    pairs = []  # [(week, code), ...]
    week = 1
    prev_token = None
    for t in tokens[start:]:
        if week > 52:
            break
        code = None
        if str(t).strip() in ("*", "＊"):
            if prev_token:
                code = _norm_rot(prev_token + " *")
            if code and pairs:
                pairs[-1] = (pairs[-1][0], code)  # overwrite previous
                continue
        if code is None:
            code = _norm_rot(t)
        if code:
            pairs.append((week, code))
            prev_token = t
            week += 1
    db.query(ScheduleAssignment).filter(
        ScheduleAssignment.resident_id == resident_id,
        ScheduleAssignment.year_id == year_id,
    ).delete()
    for w, c in pairs:
        db.add(ScheduleAssignment(resident_id=resident_id, year_id=year_id, week_number=w, rotation_code=c))
    db.commit()
    return {"ok": True, "assignments_added": len(pairs)}


def _norm_rot(val):
    """Normalize Excel rotation cell to canonical code, or None if invalid."""
    if val is None:
        return None
    s = str(val).strip().lower()
    if not s:
        return None
    return ROT_NORMALIZE.get(s)


def _import_from_schedule_sheet(content: bytes, year_id: int, db: Session) -> dict:
    """Import roster + schedule from SCHEDULE sheet. Returns {created, assignments} or -1 if no sheet."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    if "SCHEDULE" not in wb.sheetnames:
        return {"_ok": False}
    ws = wb["SCHEDULE"]

    # Clear existing residents and assignments for this year (replace import)
    db.query(ScheduleAssignment).filter(ScheduleAssignment.year_id == year_id).delete()
    db.query(Resident).filter(Resident.year_id == year_id).delete()
    db.commit()

    cohorts = {c.name: c.id for c in db.query(Cohort).filter(Cohort.year_id == year_id).all()}
    current_cohort = None
    current_pgy = None
    row_to_resident = []  # (row, name, pgy, cohort_id)
    for row in range(4, 57):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        c = ws.cell(row, 3).value
        if a and str(a).strip().startswith("Cohort"):
            current_cohort = str(a).strip()
        if b:
            ps = str(b).strip().upper().replace("-", "")
            if ps == "PGY1":
                current_pgy = "PGY1"
            elif ps == "PGY2":
                current_pgy = "PGY2"
            elif ps == "PGY3":
                current_pgy = "PGY3"
            elif ps == "TY":
                current_pgy = "TY"
        if not c or not str(c).strip() or not current_pgy:
            row_to_resident.append((row, "", "", None))
            continue
        name = str(c).strip()
        cohort_id = cohorts.get(current_cohort) if current_cohort else None
        if cohort_id:
            n_in_cohort = sum(1 for _, _, _, cid in row_to_resident if cid == cohort_id)
            if n_in_cohort >= MAX_COHORT_SIZE:
                cname = current_cohort or str(cohort_id)
                raise HTTPException(400, f"Cohort {cname} would exceed {MAX_COHORT_SIZE} residents. Max per cohort is {MAX_COHORT_SIZE}.")
        r = Resident(name=name, pgy=current_pgy, cohort_id=cohort_id, year_id=year_id)
        db.add(r)
        db.flush()
        row_to_resident.append((row, name, current_pgy, cohort_id))

    intern_per_cohort = {}
    for _, _, pgy, cid in row_to_resident:
        if cid and _is_intern(pgy):
            intern_per_cohort[cid] = intern_per_cohort.get(cid, 0) + 1
    for cid, n in intern_per_cohort.items():
        if n % 2 != 0:
            c = db.query(Cohort).filter(Cohort.id == cid).first()
            cname = c.name if c else str(cid)
            db.rollback()
            raise HTTPException(400, f"Cohort {cname} has {n} interns. Interns must be in multiples of 2 (each needs a co-intern).")
    db.commit()
    # Map row -> resident_id (residents created in row order)
    residents_by_row = {}
    all_residents = db.query(Resident).filter(Resident.year_id == year_id).order_by(Resident.id).all()
    idx = 0
    for row, name, pgy, _ in row_to_resident:
        if name and pgy and idx < len(all_residents):
            residents_by_row[row] = all_residents[idx].id
            idx += 1
    created = idx

    # Import schedule assignments: cols 4-55 = weeks 1-52
    n_assign = 0
    for row, name, pgy, _ in row_to_resident:
        if not name or row not in residents_by_row:
            continue
        rid = residents_by_row[row]
        for week in range(1, 53):
            col = 3 + week  # col 4 = week 1
            val = ws.cell(row, col).value
            code = _norm_rot(val)
            if code:
                db.add(ScheduleAssignment(
                    resident_id=rid,
                    year_id=year_id,
                    week_number=week,
                    rotation_code=code,
                ))
                n_assign += 1
    db.commit()
    return {"_ok": True, "created": created, "assignments": n_assign}


@router.post("/import")
def import_roster(year_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import roster from Excel. Supports: (1) SCHEDULE sheet rows 4-56 A=Cohort,B=PGY,C=Name; (2) any sheet with Name, PGY, Cohort columns."""
    content = file.file.read()
    year = db.query(Year).filter(Year.id == year_id).first()
    if not year:
        raise HTTPException(400, "Year not found")
    if db.query(Cohort).filter(Cohort.year_id == year_id).count() == 0:
        for i in range(1, 6):
            db.add(Cohort(year_id=year_id, name=f"Cohort {i}", clinic_weeks=[], target_intern_count=2))
        db.commit()

    schedule_err = None
    try:
        result = _import_from_schedule_sheet(content, year_id, db)
        if result.get("_ok"):
            return {"created": result["created"], "assignments": result.get("assignments", 0)}
    except Exception as e:
        schedule_err = str(e)

    # Fallback: try pandas with SCHEDULE sheet or first sheet
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        sheet_name = "SCHEDULE" if "SCHEDULE" in wb.sheetnames else wb.sheetnames[0]
        df = pd.read_excel(BytesIO(content), sheet_name=sheet_name, engine="openpyxl")
    except Exception as e:
        raise HTTPException(400, f"Invalid Excel: {e}. {f'Schedule import failed: {schedule_err}' if schedule_err else ''}")

    def col_match(pat: str):
        return next((c for c in df.columns if pat in str(c).lower()), None)
    name_col = col_match("name") or col_match("resident") or (df.columns[0] if len(df.columns) > 0 else None)
    pgy_col = col_match("pgy")
    cohort_col = next((c for c in df.columns if "cohort" in str(c).lower()), None)
    if not pgy_col or not name_col:
        raise HTTPException(
            400,
            f"Excel must have Name and PGY columns. Found: {list(df.columns)[:10]}. "
            f"{f'Schedule import error: {schedule_err}' if schedule_err else ''}"
        )

    cohorts = {c.name: c.id for c in db.query(Cohort).filter(Cohort.year_id == year_id).all()}
    existing_counts = {}
    for r in db.query(Resident).filter(Resident.year_id == year_id):
        if r.cohort_id:
            existing_counts[r.cohort_id] = existing_counts.get(r.cohort_id, 0) + 1
    batch_counts = {}
    created = 0
    for _, row in df.iterrows():
        name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
        if not name:
            continue
        pgy = str(row[pgy_col]).strip().upper().replace("-", "")
        if pgy not in ("PGY1", "PGY2", "PGY3", "TY"):
            pgy = "PGY1"
        cohort_id = None
        if cohort_col and cohort_col in row.index and pd.notna(row.get(cohort_col)):
            val = str(row[cohort_col]).strip()
            cohort_id = cohorts.get(val)
        if cohort_id:
            n = existing_counts.get(cohort_id, 0) + batch_counts.get(cohort_id, 0)
            if n >= MAX_COHORT_SIZE:
                c = db.query(Cohort).filter(Cohort.id == cohort_id).first()
                cname = c.name if c else str(cohort_id)
                raise HTTPException(400, f"Cohort {cname} would exceed {MAX_COHORT_SIZE} residents. Max per cohort is {MAX_COHORT_SIZE}.")
            batch_counts[cohort_id] = batch_counts.get(cohort_id, 0) + 1
        r = Resident(name=name, pgy=pgy, cohort_id=cohort_id, year_id=year_id)
        db.add(r)
        created += 1
    db.flush()
    intern_per_cohort = {}
    for r in db.query(Resident).filter(Resident.year_id == year_id):
        if r.cohort_id and _is_intern(r.pgy):
            intern_per_cohort[r.cohort_id] = intern_per_cohort.get(r.cohort_id, 0) + 1
    for cid, n in intern_per_cohort.items():
        if n % 2 != 0:
            c = db.query(Cohort).filter(Cohort.id == cid).first()
            cname = c.name if c else str(cid)
            db.rollback()
            raise HTTPException(400, f"Cohort {cname} has {n} interns. Interns must be in multiples of 2 (each needs a co-intern).")
    db.commit()
    return {"created": created}
