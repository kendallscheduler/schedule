"""Rollover: create next year roster from current year with placeholders."""
from typing import Optional, List, Dict
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from pydantic import BaseModel
from io import BytesIO

import openpyxl

from database import get_db
from models import (
    Resident, Year, Cohort, Week, CoverageRule, ScheduleAssignment, Completion,
)

# NF counts as FLOORS. SWING counts as NF or ICU_NIGHT (assign optimally).
_FLOOR_ROTS = ["A", "B", "C", "D", "G"]
_ICU_ROTS = ["ICU", "ICU E", "ICU N"]
_OTHER_CAT_ROTS = {
    "CLINIC": ["CLINIC", "CLINIC *"],
    "CARDIO": ["CARDIO", "CARDIO-RAM", "CARDIO-HCA"],
    "ED": ["ED"], "ID": ["ID"], "NEURO": ["NEURO"],
    "VACATION": ["VACATION"],
    "GERIATRICS": ["GERIATRICS"], "PULMONOLOGY": ["PULMONOLOGY"],
    "NEPHROLOGY": ["NEPHROLOGY"], "PALLIATIVE": ["PALLIATIVE"],
    "PAIN": ["PAIN"], "RHEUMATOLOGY": ["RHEUMATOLOGY"], "ENDOCRINOLOGY": ["ENDOCRINOLOGY"],
}


def _aggregate_assignments_to_categories(assignments: List[tuple]) -> Dict[str, int]:
    """Given [(week, rot), ...], return {category: weeks}. NF->FLOORS, SWING->NF or ICU_NIGHT (balanced)."""
    floor_cnt = sum(1 for _, rot in assignments if rot in _FLOOR_ROTS)
    nf_cnt = sum(1 for _, rot in assignments if rot == "NF")
    icun_cnt = sum(1 for _, rot in assignments if rot == "ICU N")
    swing_cnt = sum(1 for _, rot in assignments if rot == "SWING")
    x = (icun_cnt - nf_cnt + swing_cnt) // 2 if swing_cnt else 0
    swing_to_nf = max(0, min(swing_cnt, x))
    by_cat: Dict[str, int] = {}
    by_cat["FLOORS"] = floor_cnt + nf_cnt + swing_to_nf
    by_cat["ICU"] = sum(1 for _, rot in assignments if rot in _ICU_ROTS)
    by_cat["NF"] = nf_cnt + swing_to_nf
    by_cat["ICU_NIGHT"] = icun_cnt + (swing_cnt - swing_to_nf)
    by_cat["SWING"] = swing_cnt
    for _w, rot in assignments:
        for cat, rots in _OTHER_CAT_ROTS.items():
            if rot in rots:
                by_cat[cat] = by_cat.get(cat, 0) + 1
                break
    return by_cat

DATE_RANGES = [
    ("07/01", "07/06"), ("07/07", "07/13"), ("07/14", "07/20"), ("07/21", "07/27"), ("07/28", "08/03"),
    ("08/04", "08/10"), ("08/11", "08/17"), ("08/18", "08/24"), ("08/25", "08/31"), ("09/01", "09/07"),
    ("09/08", "09/14"), ("09/15", "09/21"), ("09/22", "09/28"), ("09/29", "10/05"), ("10/06", "10/12"),
    ("10/13", "10/19"), ("10/20", "10/26"), ("10/27", "11/02"), ("11/03", "11/09"), ("11/10", "11/16"),
    ("11/17", "11/23"), ("11/24", "11/30"), ("12/01", "12/07"), ("12/08", "12/14"), ("12/15", "12/21"),
    ("12/22", "12/28"), ("12/29", "01/04"), ("01/05", "01/11"), ("01/12", "01/18"), ("01/19", "01/25"),
    ("01/26", "02/01"), ("02/02", "02/08"), ("02/09", "02/15"), ("02/16", "02/22"), ("02/23", "03/01"),
    ("03/02", "03/08"), ("03/09", "03/15"), ("03/16", "03/22"), ("03/23", "03/29"), ("03/30", "04/05"),
    ("04/06", "04/12"), ("04/13", "04/19"), ("04/20", "04/26"), ("04/27", "05/03"), ("05/04", "05/10"),
    ("05/11", "05/17"), ("05/18", "05/24"), ("05/25", "05/31"), ("06/01", "06/07"), ("06/08", "06/14"),
    ("06/15", "06/21"), ("06/22", "06/30"),
]
MONTHS = ["July"]*5 + ["August"]*4 + ["September"]*4 + ["October"]*4 + ["November"]*4 + ["December"]*4 + ["January"]*4 + ["February"]*4 + ["March"]*4 + ["April"]*4 + ["May"]*4 + ["June"]*4

router = APIRouter()


class CohortsConfigItem(BaseModel):
    cohort_id: int  # 1-5, maps to Cohort 1..5
    target_interns: int


def _default_cohorts_config() -> List[CohortsConfigItem]:
    return [CohortsConfigItem(cohort_id=i, target_interns=2) for i in range(1, 6)]


class RolloverRequest(BaseModel):
    source_year_id: int
    target_year_name: str  # e.g. "2026-2027"
    target_start_date: str = ""  # e.g. "2026-07-01"; auto-derived if empty
    intern_count: int = 14
    ty_count: int = 8
    include_pgy3: bool = False  # Chief coverage: include graduating PGY3
    rebalance_seniors: bool = False
    cohorts_config: Optional[list[CohortsConfigItem]] = None


def _get_next_start_date(start: str) -> str:
    """Increment year in YYYY-MM-DD."""
    parts = start.split("-")
    if len(parts) == 3:
        y = int(parts[0]) + 1
        return f"{y}-{parts[1]}-{parts[2]}"
    return start


@router.post("/")
@router.post("/rollover")
def rollover(req: RolloverRequest, db: Session = Depends(get_db)):
    """
    Create next year roster from current year.
    - PGY1 -> PGY2, PGY2 -> PGY3, PGY3 -> exclude (unless include_pgy3), TY -> exclude
    - Create placeholder Intern 01..N and TY 01..M
    - Assign interns to cohorts per cohorts_config
    """
    source = db.query(Year).filter(Year.id == req.source_year_id).first()
    if not source:
        raise HTTPException(404, "Source year not found")

    target_start = req.target_start_date or _get_next_start_date(source.start_date or "2025-07-01")

    target = db.query(Year).filter(Year.name == req.target_year_name).first()
    if target:
        # Clear existing residents and assignments for target (allow re-rollover)
        db.query(ScheduleAssignment).filter(ScheduleAssignment.year_id == target.id).delete()
        db.query(Resident).filter(Resident.year_id == target.id).delete()
    else:
        target = Year(name=req.target_year_name, start_date=target_start)
        db.add(target)
        db.commit()
        db.refresh(target)

    # Create weeks for target if missing
    if db.query(Week).filter(Week.year_id == target.id).count() == 0:
        for i, ((s, e), m) in enumerate(zip(DATE_RANGES, MONTHS)):
            db.add(Week(year_id=target.id, week_number=i + 1, start_date=s, end_date=e, month_label=m))
        db.commit()

    # Create cohorts for target
    source_cohorts = {c.name: c for c in db.query(Cohort).filter(Cohort.year_id == source.id).all()}
    target_cohorts = {}
    for name in ["Cohort 1", "Cohort 2", "Cohort 3", "Cohort 4", "Cohort 5"]:
        tc = db.query(Cohort).filter(Cohort.year_id == target.id, Cohort.name == name).first()
        if not tc:
            sc = source_cohorts.get(name)
            weeks = list(sc.clinic_weeks) if sc and sc.clinic_weeks else []
            tc = Cohort(year_id=target.id, name=name, clinic_weeks=weeks, target_intern_count=2)
            db.add(tc)
        db.flush()
        target_cohorts[name] = tc

    db.commit()
    # Refresh to get ids
    target_cohorts = {c.name: c for c in db.query(Cohort).filter(Cohort.year_id == target.id).all()}

    # Build cohorts_config: cohort_id 1-5 -> target_interns (must be even: 2, 4, 6, 8)
    cohorts_config = req.cohorts_config or _default_cohorts_config()
    config_by_num = {}
    for c in cohorts_config:
        n = c.target_interns
        if n % 2 != 0:
            n = max(0, n - 1)
        config_by_num[c.cohort_id] = n
    cohort_order = ["Cohort 1", "Cohort 2", "Cohort 3", "Cohort 4", "Cohort 5"]

    # Promote existing residents and carry forward rotation history
    promoted = []
    resident_id_map = {}  # old_id -> new Resident
    for r in db.query(Resident).filter(Resident.year_id == source.id).all():
        new_cohort_id = None
        if r.cohort:
            cname = r.cohort.name
            if cname in target_cohorts:
                new_cohort_id = target_cohorts[cname].id

        if r.pgy == "PGY1":
            new_pgy = "PGY2"
        elif r.pgy == "PGY2":
            new_pgy = "PGY3"
        elif r.pgy == "PGY3":
            if not req.include_pgy3:
                continue
            new_pgy = "PGY3"
        else:
            continue  # TY graduates

        nr = Resident(
            name=r.name,
            pgy=new_pgy,
            cohort_id=new_cohort_id,
            prior_resident_id=r.id,
            track=r.track,
            year_id=target.id,
            constraints_json=r.constraints_json or {},
            is_placeholder=False,
        )
        db.add(nr)
        db.flush()
        resident_id_map[r.id] = nr
        promoted.append(new_pgy)

    # Copy completions + source year schedule into new residents' completions
    for old_id, new_res in resident_id_map.items():
        # Existing completions from source resident
        comps = {c.category: c.completed_weeks for c in db.query(Completion).filter(Completion.resident_id == old_id).all()}
        # Add weeks from source year's schedule
        assigns = [(a.week_number, a.rotation_code) for a in db.query(ScheduleAssignment).filter(
            ScheduleAssignment.resident_id == old_id, ScheduleAssignment.year_id == source.id
        ).all()]
        from_schedule = _aggregate_assignments_to_categories(assigns)
        for cat, weeks in from_schedule.items():
            comps[cat] = comps.get(cat, 0) + weeks
        for cat, weeks in comps.items():
            if weeks > 0:
                db.add(Completion(resident_id=new_res.id, category=cat, completed_weeks=weeks, source="rollover"))

    db.flush()

    # Count promoted per cohort (PGY3s graduate, so not included)
    from engine import MAX_COHORT_SIZE
    promoted_per_cohort = {cname: 0 for cname in cohort_order}
    for r in db.query(Resident).filter(Resident.year_id == target.id).all():
        if r.cohort_id and r.cohort:
            cname = r.cohort.name
            if cname in promoted_per_cohort:
                promoted_per_cohort[cname] += 1

    # Create placeholder PGY1 interns: cap each cohort at (MAX - promoted) to stay within limit
    intern_count = req.intern_count if req.intern_count % 2 == 0 else max(0, req.intern_count - 1)
    intern_slots = []
    for i, cname in enumerate(cohort_order, 1):
        target_n = config_by_num.get(i, 2)
        slots_available = max(0, MAX_COHORT_SIZE - promoted_per_cohort.get(cname, 0))
        n = min(target_n, slots_available)
        n = n if n % 2 == 0 else max(0, n - 1)
        for _ in range(n):
            intern_slots.append(cname)
    overflow = intern_count - len(intern_slots)
    if overflow > 0:
        for _ in range(overflow // 2):
            best = None
            for cname in cohort_order:
                current = sum(1 for s in intern_slots if s == cname)
                if promoted_per_cohort.get(cname, 0) + current + 2 <= MAX_COHORT_SIZE:
                    best = cname
                    break
            if best is None:
                db.rollback()
                raise HTTPException(
                    400,
                    "Not enough cohort capacity for all interns. PGY3s graduating free slots, but total interns still exceeds available space. "
                    "Reduce incoming intern count or adjust cohort targets.",
                )
            intern_slots.append(best)
            intern_slots.append(best)
    elif overflow < 0:
        intern_slots = intern_slots[:intern_count]

    for i in range(intern_count):
        cname = intern_slots[i]
        cohort_id = target_cohorts.get(cname).id if cname in target_cohorts else None
        nr = Resident(
            name=f"Intern {i+1:02d}",
            pgy="PGY1",
            cohort_id=cohort_id,
            year_id=target.id,
            constraints_json={"no_cardio_before_week": 7},
            is_placeholder=True,
        )
        db.add(nr)

    # Create placeholder TYs (no cohort)
    for i in range(req.ty_count):
        nr = Resident(
            name=f"TY {i+1:02d}",
            pgy="TY",
            cohort_id=None,
            year_id=target.id,
            is_placeholder=True,
        )
        db.add(nr)

    # Copy coverage rules for target
    if db.query(CoverageRule).filter(CoverageRule.year_id == target.id).count() == 0:
        for cr in db.query(CoverageRule).filter(CoverageRule.year_id == source.id).all():
            db.add(CoverageRule(
                year_id=target.id,
                pool=cr.pool,
                required_units_per_week=cr.required_units_per_week,
                seniors_per_unit=cr.seniors_per_unit,
                interns_per_unit=cr.interns_per_unit,
                optional=cr.optional,
            ))
    db.flush()
    from engine import MAX_COHORT_SIZE
    from sqlalchemy import func
    cohort_counts = db.query(Resident.cohort_id, func.count(Resident.id)).filter(
        Resident.year_id == target.id, Resident.cohort_id.isnot(None)
    ).group_by(Resident.cohort_id).all()
    for cid, n in cohort_counts:
        if n > MAX_COHORT_SIZE:
            c = db.query(Cohort).filter(Cohort.id == cid).first()
            cname = c.name if c else str(cid)
            db.rollback()
            raise HTTPException(400, f"Rollover would put {n} residents in Cohort {cname}. Max is {MAX_COHORT_SIZE}. Adjust cohort targets.")
    db.commit()
    n_res = db.query(Resident).filter(Resident.year_id == target.id).count()
    return {
        "target_year_id": target.id,
        "target_year_name": target.name,
        "promoted_count": len(promoted),
        "intern_placeholders": req.intern_count,
        "ty_placeholders": req.ty_count,
        "total_residents": n_res,
    }


@router.post("/from-excel")
def rollover_from_excel(
    file: UploadFile = File(...),
    target_year_name: str = Query("2026-2027"),
    intern_count: int = Query(14),
    ty_count: int = Query(8),
    include_pgy3: bool = Query(False),
    cohort_1_target: int = Query(4),
    cohort_2_target: int = Query(2),
    cohort_3_target: int = Query(2),
    cohort_4_target: int = Query(2),
    cohort_5_target: int = Query(2),
    db: Session = Depends(get_db),
):
    """
    Import current roster from Excel and create next year.
    Expects SCHEDULE sheet: rows 4-56, cols A=Cohort, B=PGY, C=Name.
    """
    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Invalid Excel: {e}")
    if "SCHEDULE" not in wb.sheetnames:
        raise HTTPException(400, "Excel must have SCHEDULE sheet")

    ws = wb["SCHEDULE"]
    current_cohort = None
    current_pgy = None
    roster = []  # [(name, pgy, cohort_name)]
    for row in range(4, 57):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        c = ws.cell(row, 3).value
        if a and str(a).strip().startswith("Cohort"):
            current_cohort = str(a).strip()
        if b:
            ps = str(b).strip().upper()
            if ps in ("PGY-1", "PGY1"):
                current_pgy = "PGY1"
            elif ps in ("PGY-2", "PGY2"):
                current_pgy = "PGY2"
            elif ps in ("PGY-3", "PGY3"):
                current_pgy = "PGY3"
            elif ps == "TY":
                current_pgy = "TY"
        if not c or not str(c).strip() or not current_pgy:
            continue
        name = str(c).strip()
        roster.append((name, current_pgy, current_cohort or ""))

    # Create source year in memory (we don't persist it)
    source_year = db.query(Year).filter(Year.name == "2025-2026").first()
    if not source_year:
        source_year = db.query(Year).first()
    if not source_year:
        raise HTTPException(400, "No years in DB. Run seed first.")

    target = db.query(Year).filter(Year.name == target_year_name).first()
    if target:
        db.query(ScheduleAssignment).filter(ScheduleAssignment.year_id == target.id).delete()
        db.query(Resident).filter(Resident.year_id == target.id).delete()
    else:
        target = Year(name=target_year_name, start_date="2026-07-01")
        db.add(target)
        db.commit()
        db.refresh(target)

    if db.query(Week).filter(Week.year_id == target.id).count() == 0:
        for i, ((s, e), m) in enumerate(zip(DATE_RANGES, MONTHS)):
            db.add(Week(year_id=target.id, week_number=i + 1, start_date=s, end_date=e, month_label=m))
        db.commit()

    target_cohorts = {}
    for name in ["Cohort 1", "Cohort 2", "Cohort 3", "Cohort 4", "Cohort 5"]:
        tc = db.query(Cohort).filter(Cohort.year_id == target.id, Cohort.name == name).first()
        if not tc:
            tc = Cohort(year_id=target.id, name=name, clinic_weeks=[], target_intern_count=2)
            db.add(tc)
    db.commit()
    target_cohorts = {c.name: c for c in db.query(Cohort).filter(Cohort.year_id == target.id).all()}

    def _even(n: int) -> int:
        return n if n % 2 == 0 else max(0, n - 1)
    intern_count = _even(intern_count)
    cohorts_config = [
        CohortsConfigItem(cohort_id=1, target_interns=_even(cohort_1_target)),
        CohortsConfigItem(cohort_id=2, target_interns=_even(cohort_2_target)),
        CohortsConfigItem(cohort_id=3, target_interns=_even(cohort_3_target)),
        CohortsConfigItem(cohort_id=4, target_interns=_even(cohort_4_target)),
        CohortsConfigItem(cohort_id=5, target_interns=_even(cohort_5_target)),
    ]
    config_by_num = {c.cohort_id: c.target_interns for c in cohorts_config}
    cohort_order = ["Cohort 1", "Cohort 2", "Cohort 3", "Cohort 4", "Cohort 5"]

    promoted = 0
    for name, pgy, cohort_name in roster:
        if pgy == "TY":
            continue
        if pgy == "PGY3" and not include_pgy3:
            continue
        if pgy == "PGY1":
            new_pgy = "PGY2"
        elif pgy == "PGY2":
            new_pgy = "PGY3"
        else:
            new_pgy = "PGY3"
        cohort_id = target_cohorts.get(cohort_name).id if cohort_name in target_cohorts else None
        nr = Resident(name=name, pgy=new_pgy, cohort_id=cohort_id, year_id=target.id, is_placeholder=False)
        db.add(nr)
        promoted += 1

    db.flush()
    from engine import MAX_COHORT_SIZE as MAX_COHORT
    promoted_per_cohort = {cname: 0 for cname in cohort_order}
    for r in db.query(Resident).filter(Resident.year_id == target.id).all():
        if r.cohort_id and r.cohort and r.cohort.name in promoted_per_cohort:
            promoted_per_cohort[r.cohort.name] += 1
    intern_slots = []
    for i, cname in enumerate(cohort_order, 1):
        target_n = config_by_num.get(i, 2)
        slots_available = max(0, MAX_COHORT - promoted_per_cohort.get(cname, 0))
        n = min(target_n, slots_available)
        n = _even(n)
        for _ in range(n):
            intern_slots.append(cname)
    overflow = intern_count - len(intern_slots)
    if overflow > 0:
        for _ in range(overflow // 2):
            best = None
            for cname in cohort_order:
                current = sum(1 for s in intern_slots if s == cname)
                if promoted_per_cohort.get(cname, 0) + current + 2 <= MAX_COHORT:
                    best = cname
                    break
            if best is None:
                db.rollback()
                raise HTTPException(
                    400,
                    "Not enough cohort capacity for all interns. Reduce incoming intern count or cohort targets.",
                )
            intern_slots.append(best)
            intern_slots.append(best)
    elif overflow < 0:
        intern_slots = intern_slots[:intern_count]

    for i in range(intern_count):
        cname = intern_slots[i]
        cohort_id = target_cohorts.get(cname).id if cname in target_cohorts else None
        nr = Resident(name=f"Intern {i+1:02d}", pgy="PGY1", cohort_id=cohort_id, year_id=target.id, is_placeholder=True,
                     constraints_json={"no_cardio_before_week": 7})
        db.add(nr)

    for i in range(ty_count):
        nr = Resident(name=f"TY {i+1:02d}", pgy="TY", cohort_id=None, year_id=target.id, is_placeholder=True)
        db.add(nr)

    if db.query(CoverageRule).filter(CoverageRule.year_id == target.id).count() == 0:
        for cr in db.query(CoverageRule).filter(CoverageRule.year_id == source_year.id).all():
            db.add(CoverageRule(year_id=target.id, pool=cr.pool, required_units_per_week=cr.required_units_per_week,
                               seniors_per_unit=cr.seniors_per_unit, interns_per_unit=cr.interns_per_unit, optional=cr.optional))
    db.flush()
    from sqlalchemy import func
    cohort_counts = db.query(Resident.cohort_id, func.count(Resident.id)).filter(
        Resident.year_id == target.id, Resident.cohort_id.isnot(None)
    ).group_by(Resident.cohort_id).all()
    for cid, n in cohort_counts:
        if n > MAX_COHORT:
            c = db.query(Cohort).filter(Cohort.id == cid).first()
            cname = c.name if c else str(cid)
            db.rollback()
            raise HTTPException(400, f"Rollover would put {n} residents in Cohort {cname}. Max is {MAX_COHORT}. Adjust cohort targets.")
    db.commit()
    n_res = db.query(Resident).filter(Resident.year_id == target.id).count()
    return {"target_year_id": target.id, "target_year_name": target.name, "promoted_count": promoted,
            "intern_placeholders": intern_count, "ty_placeholders": ty_count, "total_residents": n_res}
