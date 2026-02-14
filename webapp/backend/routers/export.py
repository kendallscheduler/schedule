"""Export schedule to Excel matching existing master schedule layout."""
import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db
from models import Resident, ScheduleAssignment, Year, Cohort

router = APIRouter()


@router.get("/excel")
def export_excel(year_id: int, db: Session = Depends(get_db)):
    try:
        """Export to Excel with grid: rows=residents, cols=weeks 1-52."""
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import traceback

        residents = (
            db.query(Resident)
            .filter(Resident.year_id == year_id)
            .order_by(Resident.cohort_id, Resident.pgy, Resident.name)
            .all()
        )
        
        assignments = {}
        for a in db.query(ScheduleAssignment).filter(ScheduleAssignment.year_id == year_id).all():
            assignments.setdefault(a.resident_id, {})[a.week_number] = a.rotation_code

        # Blocks definition: 13 blocks of 4 weeks each
        BLOCKS = [
            {"label": "Block 1", "colspan": 4},
            {"label": "Block 2", "colspan": 4},
            {"label": "Block 3", "colspan": 4},
            {"label": "Block 4", "colspan": 4},
            {"label": "Block 5", "colspan": 4},
            {"label": "Block 6", "colspan": 4},
            {"label": "Block 7", "colspan": 4},
            {"label": "Block 8", "colspan": 4},
            {"label": "Block 9", "colspan": 4},
            {"label": "Block 10", "colspan": 4},
            {"label": "Block 11", "colspan": 4},
            {"label": "Block 12", "colspan": 4},
            {"label": "Block 13", "colspan": 4},
        ]
        
        block_edges = set()
        current_sum = 0
        for b in BLOCKS:
            current_sum += b["colspan"]
            block_edges.add(current_sum)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Master Schedule"

        # Refined Styles
        border_color = "CBD5E1"
        thick_color = "64748B"
        thin_side = Side(border_style="thin", color=border_color)
        thick_side = Side(border_style="medium", color=thick_color)
        
        center_align = Alignment(horizontal="center", vertical="center")
        header_font = Font(bold=True, size=11, name='Calibri')
        cell_font = Font(size=9, name='Calibri')
        
        header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        block_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        block_font = Font(bold=True, size=11, color="FFFFFF", name='Calibri')

        # Colors (ARGB) - Fix: openpyxl expects hex strings without #, and alpha prefix (FF)
        COLORS = {
            "A": "FF86EFAC", "B": "FF86EFAC", "C": "FF86EFAC", "D": "FF86EFAC", "G": "FF86EFAC",
            "ICU": "FF7DD3FC", "ICU N": "FF38BDF8",
            "CLINIC": "FF93C5FD", "CLINIC *": "FF93C5FD", "ED": "FF93C5FD", "GEN SURG": "FF93C5FD", "TY CLINIC": "FF93C5FD",
            "VACATION": "FFFCA5A5", "NF": "FFE2E8F0", "SWING": "FFC4B5FD",
            "CARDIO": "FFFDBA74", 
            "ID": "FFFEF08A", "NEURO": "FFFEF08A", "GERIATRICS": "FFFEF08A", "ELECTIVE": "FFFEF08A", "ANESTHESIA": "FFFEF08A",
        }
        
        # Set headers (Row 1 & 2)
        # Resident, PGY, Track merged across 2 rows
        for col_idx, val in enumerate(["Resident", "PGY", "Track"], 1):
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            c = ws.cell(1, col_idx, val)
            c.font = header_font
            c.alignment = center_align
            c.fill = header_fill
            # Borders for merged info cells
            for r_i in [1, 2]:
                ws.cell(r_i, col_idx).border = Border(top=thin_side if r_i==1 else None, bottom=thin_side if r_i==2 else None, left=thin_side, right=thick_side if col_idx==3 else thin_side)

        # Draw Block Headers (Row 1)
        curr_col = 4
        for b in BLOCKS:
            ws.merge_cells(start_row=1, start_column=curr_col, end_row=1, end_column=curr_col + b["colspan"] - 1)
            cell = ws.cell(1, curr_col)
            cell.value = b["label"]
            cell.font = block_font
            cell.alignment = center_align
            cell.fill = block_fill
            
            # Apply borders to the entire merged block range header
            for c_idx in range(curr_col, curr_col + b["colspan"]):
                top_cell = ws.cell(1, c_idx)
                r_s = thick_side if c_idx == curr_col + b["colspan"] - 1 else thin_side
                top_cell.border = Border(top=thin_side, bottom=thin_side, left=thin_side if c_idx == curr_col else None, right=r_s)
                
            curr_col += b["colspan"]

        # Week Numbers Row (Row 2)
        for w in range(1, 53):
            col = w + 3
            c = ws.cell(2, col, w)
            c.font = Font(bold=True, size=9)
            c.alignment = center_align
            r_side = thick_side if w in block_edges else thin_side
            c.border = Border(bottom=thin_side, right=r_side)
            c.fill = header_fill

        # Resident Rows (Row 3 onwards)
        row_idx = 3
        for r in residents:
            # Info columns
            ws.cell(row_idx, 1, r.name).font = Font(bold=True, size=10)
            ws.cell(row_idx, 2, r.pgy).alignment = center_align
            ws.cell(row_idx, 3, r.cohort.name if r.cohort else ("TY" if r.pgy == "TY" else "")).alignment = center_align
            
            for i in range(1, 4):
                ws.cell(row_idx, i).border = Border(bottom=thin_side, left=thin_side if i==1 else None, right=thick_side if i==3 else thin_side)
                ws.cell(row_idx, i).font = Font(size=10)

            # Assignment columns
            for w in range(1, 53):
                code = assignments.get(r.id, {}).get(w, "")
                # Map rotations to display labels matching web UI
                display_label = {
                    "CLINIC *": "CLINIC*",
                    "GERIATRICS": "GERI",
                    "ELECTIVE": "ELECT",
                    "GEN SURG": "SURG",
                    "TY CLINIC": "TY CL"
                }.get(code, code)
                
                cell = ws.cell(row_idx, w + 3, display_label)
                cell.alignment = center_align
                cell.font = cell_font
                
                # Apply thick border for block edges
                r_side = thick_side if w in block_edges else thin_side
                cell.border = Border(bottom=thin_side, right=r_side)
                
                # Apply colors
                if code in COLORS:
                    cell.fill = PatternFill(start_color=COLORS[code], end_color=COLORS[code], fill_type="solid")

            ws.row_dimensions[row_idx].height = 16 # Add some row padding
            row_idx += 1

        # Freeze panes
        ws.freeze_panes = "D3"
        
        # Sizing
        ws.column_dimensions["A"].width = 28 # Wider for names
        ws.column_dimensions["B"].width = 7
        ws.column_dimensions["C"].width = 12
        for w in range(1, 53):
            ws.column_dimensions[get_column_letter(w + 3)].width = 6.5 # Wider for labels

        # Totals / Summary Section
        start_totals_col = 56
        totals_headers = ["FLOORS", "ICU", "ICU N", "NF", "SWING", "CLINIC", "ED", "CARDIO", "ELECTIVE", "VACATION"]
        for i, h in enumerate(totals_headers):
            c = ws.cell(2, start_totals_col + i, h)
            c.font = header_font
            c.alignment = center_align
            c.fill = header_fill
            c.border = Border(bottom=thin_side, right=thin_side, top=thin_side)
            ws.column_dimensions[get_column_letter(start_totals_col + i)].width = 11

        for r_row in range(3, row_idx):
            # FLOORS
            ws.cell(r_row, start_totals_col, f'=COUNTIF(D{r_row}:BC{r_row},"A")+COUNTIF(D{r_row}:BC{r_row},"B")+COUNTIF(D{r_row}:BC{r_row},"C")+COUNTIF(D{r_row}:BC{r_row},"D")+COUNTIF(D{r_row}:BC{r_row},"G")')
            # ICU
            ws.cell(r_row, start_totals_col+1, f'=COUNTIF(D{r_row}:BC{r_row},"ICU")')
            # ICU N
            ws.cell(r_row, start_totals_col+2, f'=COUNTIF(D{r_row}:BC{r_row},"ICU N")')
            # NF
            ws.cell(r_row, start_totals_col+3, f'=COUNTIF(D{r_row}:BC{r_row},"NF")')
            # SWING
            ws.cell(r_row, start_totals_col+4, f'=COUNTIF(D{r_row}:BC{r_row},"SWING")')
            # CLINIC
            ws.cell(r_row, start_totals_col+5, f'=COUNTIF(D{r_row}:BC{r_row},"CLINIC")+COUNTIF(D{r_row}:BC{r_row},"CLINIC*")+COUNTIF(D{r_row}:BC{r_row},"TY CL")')
            # ED
            ws.cell(r_row, start_totals_col+6, f'=COUNTIF(D{r_row}:BC{r_row},"ED")')
            # CARDIO
            ws.cell(r_row, start_totals_col+7, f'=COUNTIF(D{r_row}:BC{r_row},"CARDIO")')
            # ELECTIVE
            ws.cell(r_row, start_totals_col+8, f'=COUNTIF(D{r_row}:BC{r_row},"ELECT")+COUNTIF(D{r_row}:BC{r_row},"ID")+COUNTIF(D{r_row}:BC{r_row},"NEURO")+COUNTIF(D{r_row}:BC{r_row},"GERI")+COUNTIF(D{r_row}:BC{r_row},"GEN SURG")+COUNTIF(D{r_row}:BC{r_row},"ANESTHESIA")')
            # VACATION
            ws.cell(r_row, start_totals_col+9, f'=COUNTIF(D{r_row}:BC{r_row},"VACATION")')
            
            for i in range(10):
                cell = ws.cell(r_row, start_totals_col + i)
                cell.border = Border(bottom=thin_side, right=thin_side)
                cell.alignment = center_align

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=master_schedule.xlsx"},
        )
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=traceback.format_exc())
