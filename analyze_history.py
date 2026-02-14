
import openpyxl
from scheduler.models import CODE_TO_COUNTER

def analyze_existing_schedule(path):
    print(f"Analyzing Existing Schedule: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    if "SCHEDULE" not in wb.sheetnames:
        print("No SCHEDULE sheet found.")
        return
    ws = wb["SCHEDULE"]
    
    # 1. Count Residents
    resident_rows = []
    for r in range(4, 100):
        name = ws.cell(r, 3).value
        cat = ws.cell(r, 1).value
        # If we see "Cohort" or a name, it's a valid area.
        # If we see empty for too long, stop.
        if name and str(name).strip():
            resident_rows.append(r)
        elif r > 60: # Rough limit for residents
            break
            
    print(f"Total residents detected: {len(resident_rows)}")
    
    # 2. Check for existing data
    data_points = 0
    cat_counts_per_week = {} # week -> cat -> count
    
    for col in range(4, 56): # D..BC
        week = col - 3
        cat_counts_per_week[week] = {}
        for row in resident_rows:
            val = ws.cell(row, col).value
            if val:
                data_points += 1
                cat = CODE_TO_COUNTER.get(str(val).strip().upper())
                if not cat:
                    # Try matching substring
                    for k, v in CODE_TO_COUNTER.items():
                        if k in str(val).upper():
                            cat = v
                            break
                
                if cat:
                    cat_counts_per_week[week][cat] = cat_counts_per_week[week].get(cat, 0) + 1

    print(f"Total assignment data points: {data_points}")
    
    if data_points == 0:
        print("This workbook is a template (empty grid).")
        return

    # 3. Analyze weekly capacity from the data
    print("\nWeekly Capacity (Avg per week from existing data):")
    cats = ["FLOORS", "ICU", "CLINIC", "ED", "VACATION"]
    for cat in cats:
        counts = [cat_counts_per_week[w].get(cat, 0) for w in range(1, 53)]
        avg = sum(counts) / len(counts)
        max_c = max(counts)
        print(f"  {cat:10}: Avg={avg:4.1f}, Max={max_c}")

if __name__ == "__main__":
    analyze_existing_schedule("2025 Updated Schedule - Copy 2026.xlsx")
