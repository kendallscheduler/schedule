
import pandas as pd
import openpyxl
from scheduler.models import CODE_TO_COUNTER

def check_output_quality(path):
    print(f"Checking schedule quality: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["SCHEDULE"]
    
    # Read residents (names in col C, rows 4..56)
    residents = []
    for row in range(4, 57):
        name = ws.cell(row, 3).value
        pgy = ws.cell(row, 2).value
        if name:
            residents.append({"name": name, "row": row, "pgy": pgy})
            
    # Categories to check
    categories = ["FLOORS", "ICU", "CLINIC", "ED", "CARDIO", "VACATION"]
    
    # Target requirements (mapped for simplicity)
    targets = {
        "PGY1": {"FLOORS": 20, "ICU": 8, "CLINIC": 14, "ED": 2},
        "PGY2": {"FLOORS": 16, "ICU": 8, "CLINIC": 14, "ED": 2},
        "PGY3": {"FLOORS": 16, "ICU": 0, "CLINIC": 14, "ED": 0},
        "TY":   {"FLOORS": 20, "ICU": 8, "CLINIC": 4,  "ED": 2},
    }
    
    total_deficit = 0
    total_reqs = 0
    
    # Manually count based on rotation codes in D..BC (4..55)
    for res in residents:
        counts = {cat: 0 for cat in categories}
        for col in range(4, 56):
            code = ws.cell(res["row"], col).value
            if code in CODE_TO_COUNTER:
                cat = CODE_TO_COUNTER[code]
                if cat in counts:
                    counts[cat] += 1
        
        pgy = str(res["pgy"])
        if "TY" in pgy: pgy = "TY"
        elif "1" in pgy: pgy = "PGY1"
        elif "2" in pgy: pgy = "PGY2"
        elif "3" in pgy: pgy = "PGY3"
        
        if pgy in targets:
            for cat, target in targets[pgy].items():
                if target > 0:
                    total_reqs += target
                    deficit = max(0, target - counts[cat])
                    total_deficit += deficit
                    if deficit > 0:
                        # Minor printing
                        pass

    print(f"Overall Requirement Compliance:")
    print(f"  Total target weeks: {total_reqs}")
    print(f"  Total deficit weeks: {total_deficit}")
    print(f"  Compliance rate: {((total_reqs - total_deficit) / total_reqs)*100:.1f}%")

if __name__ == "__main__":
    check_output_quality("2026 Master Schedule - Auto.xlsx")
