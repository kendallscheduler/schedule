#!/usr/bin/env python3
"""
Create an UNSCHEDULED master schedule workbook by copying the template
and clearing assignment cells (D:BC for resident rows 4-56).
Preserves all formatting, conditional formatting, formulas, and CATEGORY sheet.
"""

import argparse
import shutil
from pathlib import Path
from typing import Optional

import openpyxl


def create_unscheduled_template(
    template_path: str,
    output_path: Optional[str] = None,
) -> str:
    """
    Duplicate the template workbook and clear weekly assignment cells.
    
    Args:
        template_path: Path to source template .xlsx
        output_path: Path for output (default: template_dir/2026 Unscheduled Master Schedule.xlsx)
    
    Returns:
        Path to the created output file.
    """
    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    
    if output_path is None:
        output_path = template.parent / "2026 Unscheduled Master Schedule.xlsx"
    else:
        output_path = Path(output_path)
    
    # 1) Copy byte-for-byte to preserve everything
    shutil.copy2(template, output_path)
    
    # 2) Open with openpyxl (data_only=False to preserve formulas)
    wb = openpyxl.load_workbook(output_path, data_only=False)
    
    if "SCHEDULE" not in wb.sheetnames:
        raise ValueError("Workbook must have a 'SCHEDULE' sheet")
    
    ws = wb["SCHEDULE"]
    
    # 3) Clear assignment cells: rows 4..56, columns D(4)..BC(55)
    for row in range(4, 57):
        if ws.cell(row=row, column=3).value:  # resident name present
            for col in range(4, 56):  # D through BC (52 weeks)
                ws.cell(row=row, column=col).value = None
    
    wb.save(output_path)
    return str(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Create unscheduled master schedule from template"
    )
    parser.add_argument(
        "--template",
        default="2025 Updated Schedule - Copy 2026.xlsx",
        help="Path to template workbook",
    )
    parser.add_argument(
        "--out",
        default=None,
        help="Output path (default: 2026 Unscheduled Master Schedule.xlsx in same dir)",
    )
    args = parser.parse_args()
    
    out = create_unscheduled_template(args.template, args.out)
    print(f"Created: {out}")


if __name__ == "__main__":
    main()
