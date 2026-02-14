
import openpyxl

def count_pgys(path):
    print(f"Counting PGYs in: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["SCHEDULE"]
    
    counts = {}
    for r in range(4, 100):
        name = ws.cell(r, 3).value
        pgy = ws.cell(r, 2).value
        # If name is empty, skip. If we hit enough empties, stop.
        if name:
            p = str(pgy).strip() if pgy else "None"
            counts[p] = counts.get(p, 0) + 1
        elif r > 60:
            break
            
    print(f"PGY counts: {counts}")

if __name__ == "__main__":
    count_pgys("2025 Updated Schedule - Copy 2026.xlsx")
