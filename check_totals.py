
import openpyxl

def check_completed_totals(path):
    print(f"Checking Completed Totals in: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["SCHEDULE"]
    
    # 59 is BG (FLOORS)
    # 57 is BE (CLINIC)
    # 60 is BH (ICU)
    
    print("\nSample Totals from Historic Grid:")
    print(f"{'Name':20} | {'PGY':5} | {'FLOORS':6} | {'CLINIC':6} | {'ICU':6}")
    print("-" * 55)
    
    for r in range(4, 15):
        name = ws.cell(r, 3).value
        pgy = ws.cell(r, 2).value
        # Check column headers to be sure
        # BD=56, BE=57, BF=58, BG=59, BH=60...
        floors = ws.cell(r, 59).value
        clinic = ws.cell(r, 57).value
        icu = ws.cell(r, 60).value
        
        print(f"{str(name):20} | {str(pgy):5} | {str(floors):6} | {str(clinic):6} | {str(icu):6}")

if __name__ == "__main__":
    check_completed_totals("2025 Updated Schedule - Copy 2026.xlsx")
