
import openpyxl

def inspect_grid_details(path):
    print(f"Inspecting Grid Details: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["SCHEDULE"]
    
    unique_codes = set()
    week1_data = []
    
    for r in range(4, 57):
        name = ws.cell(r, 3).value
        if not name: continue
        
        row_codes = []
        for c in range(4, 56):
            val = ws.cell(r, c).value
            if val:
                unique_codes.add(str(val).strip())
                if c == 4: # Week 1
                    week1_data.append((name, str(val).strip()))
                    
    print("\nUnique Codes found in grid:")
    print(sorted(list(unique_codes)))
    
    print("\nWeek 1 assignments:")
    for name, code in week1_data:
        print(f"  {name:20}: {code}")

if __name__ == "__main__":
    inspect_grid_details("2025 Updated Schedule - Copy 2026.xlsx")
