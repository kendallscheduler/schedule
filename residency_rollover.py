
import openpyxl
import shutil
from pathlib import Path

def rollover_residency(input_path, output_path):
    print(f"Rolling over residency from {input_path} to {output_path}")
    shutil.copy2(input_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["SCHEDULE"]
    
    # 1. Structure discovery
    # We'll map residents by their 2025 slot
    current_residents = {}
    for r in range(4, 57):
        name = ws.cell(r, 3).value
        # We need to find the PGY value even if merged
        pgy = None
        for row_up in range(r, 3, -1):
            val = ws.cell(row_up, 2).value
            if val in ["PGY-1", "PGY-2", "PGY-3", "TY"]:
                pgy = val
                break
        if name and pgy:
            current_residents[r] = {"name": name, "pgy": pgy}

    # 2. Define target slots for 2026-2027
    # PGY-1 (2025) -> PGY-2 (2026)
    # PGY-2 (2025) -> PGY-3 (2026)
    # PGY-3 (2025) -> Graduate
    # TY (2025) -> Graduate
    
    # Let's find the slot ranges in the template
    sections = {} # pgy_label -> list of rows
    curr_sec = None
    for r in range(4, 57):
        val = ws.cell(r, 2).value
        if val in ["PGY-1", "PGY-2", "PGY-3", "TY"]:
            curr_sec = val
        if curr_sec:
            sections.setdefault(curr_sec, []).append(r)

    # 3. Create the new name mapping
    new_names = {r: None for r in range(4, 57)} # row -> name
    
    # Move PGY-1s to PGY-2 slots
    pgy1_residents = [res for row, res in current_residents.items() if res["pgy"] == "PGY-1"]
    pgy2_slots = sections.get("PGY-2", [])
    for i, res in enumerate(pgy1_residents):
        if i < len(pgy2_slots):
            new_names[pgy2_slots[i]] = res["name"]

    # Move PGY-2s to PGY-3 slots
    pgy2_residents = [res for row, res in current_residents.items() if res["pgy"] == "PGY-2"]
    pgy3_slots = sections.get("PGY-3", [])
    for i, res in enumerate(pgy2_residents):
        if i < len(pgy3_slots):
            new_names[pgy3_slots[i]] = res["name"]

    # PGY-1 slots and TY slots are now EMPTY for "New Upcoming Interns"
    pgy1_slots = sections.get("PGY-1", [])
    for r in pgy1_slots:
        new_names[r] = f"[New PGY-1 Intern {r}]"
    
    ty_slots = sections.get("TY", [])
    for r in ty_slots:
        new_names[r] = f"[New TY Intern {r}]"

    # 4. Apply changes and clear labels/grid
    # Update PGY labels: PGY-1->PGY-2, PGY-2->PGY-3, PGY-3->PGY-3? No, PGY-3 slots stay PGY-3.
    # The sections don't change, just the people move.
    
    for r in range(4, 57):
        # Update name
        ws.cell(r, 3).value = new_names[r]
        # Clear grid
        for c in range(4, 56):
            ws.cell(r, c).value = None
            
    wb.save(output_path)
    print("Rollover complete.")

if __name__ == "__main__":
    rollover_residency("2025 Updated Schedule - Copy 2026.xlsx", "2026-2027 Master Schedule.xlsx")
